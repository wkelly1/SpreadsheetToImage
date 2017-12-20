#Will Kelly
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
def columnGenerator(width, pixels):
    #print("Length of pixels =", len(pixels))

    position = []
    if len(pixels) % width == 0:
        height = len(pixels) / width
        height = int(height) + 1

    else:
        height = 2

    for row in range(1, height):
        print(row)
        length = 1
        lengthCompleted = False
        if width >= 26:
            for i in range(65, 91):
                cellPosition = chr(i) + str(row)
                #print(cellPosition)
                position.append(cellPosition)
                length = length + 1

            while lengthCompleted == False:
                for elmt in range(65, 91):
                    for i in range(65, 91):

                        if length <= width:
                            cellPosition = chr(elmt) + chr(i) + str(row)
                            position.append(cellPosition)
                            length = length + 1
                            #print(cellPosition)
                        else:
                            lengthCompleted = True

        else:

            while lengthCompleted == False:
                for i in range(65, 91):
                    if length <= width:
                        cellPosition = chr(i) + str(row)
                        #print(cellPosition)
                        position.append(cellPosition)
                        length = length + 1
                    else:
                        lengthCompleted = True

    return (position)

def pixelStringSplit(pixels):
    listPixels = []
    pixels = ''.join(pixels + ", (IGNORETHISNOW)")
    #print(pixels)
    while len(pixels) != 15:
        listPixels.append(pixels[pixels.index("("):pixels.index(")")+1])
        pixels = pixels[pixels.index(","):]
        pixels = pixels[pixels.index("("):]
        #print(pixels)

    return listPixels

def rgbToHex(rgb):
    rgb1 = hex(int(rgb[rgb.index("(")+1:rgb.index(",")]))
    rgb = rgb[rgb.index(" ")+1:]
    rgb2 = hex(int(rgb[:rgb.index(",")]))
    rgb = rgb[rgb.index(" ")+1:]
    rgb3 = hex(int(rgb[:rgb.index(")")]))
    rgb1 = rgb1[2:]
    rgb2 = rgb2[2:]
    rgb3 = rgb3[2:]

    if len(rgb1) <= 1:
        rgb1 = "0" + rgb1
    if len(rgb2) <= 1:
        rgb2 = "0" + rgb2
    if len(rgb3) <= 1:
        rgb3 = "0" + rgb3

    hexColour = "00" + rgb1 + rgb2 + rgb3
    #print("hex colour = "+hexColour)
    return hexColour


image = "dog.jpg" #Type the file name in here

im = Image.open(image)
pixels = str(list(im.getdata()))

listPixels = pixelStringSplit(pixels)
print(listPixels)

width, height = im.size

columns = columnGenerator(width, listPixels)
wb = Workbook()

# grab the active worksheet
ws = wb.active

number = 0
for i in listPixels:
    print(i)
    #print(number)

    #ws[columns[number]] = int(i[i.index("(")+1:i.index(",")])
    ws[columns[number]] = i


    hexColour = str(rgbToHex(i))

    colour = PatternFill(start_color=hexColour,
                         end_color=hexColour,
                         fill_type='solid')

    ws[columns[number]].fill = colour
    number = number + 1
# Save the file
wb.save("sample.xlsx")